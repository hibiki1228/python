import openpyxl as excel
import datetime

weekname = ["月","火","水","木","金","土","日"]

wb = excel.Workbook()
ws = wb.active

# 今年の10月1日を得る
now = datetime.datetime.now()
tm = datetime.date(now.year, 10, 1)

# 366日分を繰り返してセルに入力
for i in range(1, 367):
    ws.cell(column=1, row=i, value=tm.year)
    ws.cell(column=2, row=i, value=tm.month)
    ws.cell(column=3, row=i, value=tm.day)
    ws.cell(column=4, row=i, value=weekname[tm.weekday()])
    tm = tm + datetime.timedelta(days=1)

wb.save("cal.xlsx")