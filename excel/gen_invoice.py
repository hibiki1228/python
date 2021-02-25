import openpyxl, pprint
from datetime import datetime

file_list = "list.xlsx"
file_invoice = "invoice.xlsx"
file_delivery = "delivery_slip.xlsx"
file_out_iv = "out_invoice.xlsx"
file_out_ds = "out_delivery_slip.xlsx"

wb = openpyxl.load_workbook(file_list, data_only=True)
ws = wb["sheet1"]
name = ws["A1"].value
list_data = ws["A3:F10"]

wb_iv = openpyxl.load_workbook(file_invoice)
ws_iv = wb_iv.active
wb_ds = openpyxl.load_workbook(file_delivery)
ws_ds = wb_ds.active

cdate = datetime.now().strftime("%Y/%m/%d")
ws_iv["A3"].value = name
ws_iv["F2"].value = cdate
ws_ds["A3"].value = name
ws_ds["F2"].value = cdate

for y, row in enumerate(list_data):
    for x, cell in enumerate(row):
        if (cell is None) or (cell.value is None): continue
        v = cell.value
        ws_iv.cell(row=12+y+1, column=0+x+1, value=v)
        ws_ds.cell(row=12+y+1, column=0+x+1, value=v)

wb_iv.save(file_out_iv)
wb_ds.save(file_out_ds)
print("ok")
