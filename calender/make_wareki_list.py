# ライブラリの取り込み
import wareki
import openpyxl as excel

# 新規ワークブックを作ってシートを得る
wb = excel.Workbook()
ws = wb.active
# 100年分の背歴和暦対応表を作る
ws["A1"] = "西暦"
ws["B1"] = "和暦"
start_y = 1930
for i in range(100):
    sei = start_y + i
    wa = wareki.seireki_wareki(sei)
    ws.cell(column=1, row=(2+i), value=str(sei)+"年")
    ws.cell(column=2, row=(2+i), value=wa)
    print(sei, "=", wa)
# ファイルを保存
wb.save("wareki.xlsx")